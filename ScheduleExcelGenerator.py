import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime, timedelta


class ScheduleExcelGenerator:
    def __init__(self):
        self.week_days = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]
        self.today = datetime.today()

        # Определяем начало текущей и следующей недели
        start_of_week = self.today - timedelta(days=self.today.weekday())  # Понедельник этой недели
        next_week_start = start_of_week + timedelta(weeks=1)  # Понедельник следующей недели

        self.current_week_start = start_of_week
        self.next_week_start = next_week_start

        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = "Расписание"

        # Устанавливаем ширину столбцов
        for col in range(1, 8):  # Учитываем все 7 столбцов
            self.sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16

    def create_schedule(self):
        """Создает Excel файл с расписанием для двух недель, с отступом для второй недели."""
        headers = ["Дата", "День недели", "Статус (Рабочий/Выходной)", "Начало смены", "Начало перерыва",
                   "Конец перерыва", "Конец смены"]
        self.sheet.append(headers)

        # Заполняем расписание для текущей недели
        for i in range(7):
            day = self.current_week_start + timedelta(days=i)
            day_name = self.week_days[i]
            status = "Рабочий" if day_name not in ["Среда", "Воскресенье"] else "Выходной"

            row = [
                day.strftime("%Y-%m-%d"),
                day_name,
                status,
                "09:00" if status == "Рабочий" else "",
                "12:00" if status == "Рабочий" else "",
                "13:00" if status == "Рабочий" else "",
                "18:00" if status == "Рабочий" else ""
            ]
            self.sheet.append(row)

        # Добавляем пустую строку перед второй неделей
        self.sheet.append([""] * 7)

        # Заполняем расписание для следующей недели
        for i in range(7):
            day = self.next_week_start + timedelta(days=i)
            day_name = self.week_days[i]
            status = "Рабочий" if day_name not in ["Среда", "Воскресенье"] else "Выходной"

            row = [
                day.strftime("%Y-%m-%d"),
                day_name,
                status,
                "09:00" if status == "Рабочий" else "",
                "12:00" if status == "Рабочий" else "",
                "13:00" if status == "Рабочий" else "",
                "18:00" if status == "Рабочий" else ""
            ]
            self.sheet.append(row)


        # Центрируем текст в ячейках
        for row in self.sheet.iter_rows(min_row=2, max_row=self.sheet.max_row, min_col=1, max_col=7):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    def save(self, filename="schedule.xlsx"):
        """Сохраняет файл Excel."""
        self.workbook.save(filename)

    def get_schedule(self):
        """Возвращает путь к созданному файлу."""
        return "schedule.xlsx"


# Создаем и сохраняем расписание
generator = ScheduleExcelGenerator()
generator.create_schedule()
generator.save()
