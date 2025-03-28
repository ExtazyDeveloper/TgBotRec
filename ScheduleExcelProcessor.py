import openpyxl
import logging
from telegram import Update
from telegram.ext import CallbackContext, ConversationHandler
from datetime import datetime

# Настройка логирования
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ScheduleExcelProcessor:
    WAITING_FOR_FILE = range(1)  # Константа для состояния ожидания файла

    def __init__(self, database):
        self.db = database  # Объект базы данных для записи данных
        self.week_days = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота", "Воскресенье"]

    async def start_get_schedule(self, update: Update, context: CallbackContext):
        """Стартовый метод, который запускает процесс ожидания файла."""
        await update.message.reply_text("Пожалуйста, отправьте файл с расписанием в формате Excel.")
        return self.WAITING_FOR_FILE

    async def get_schedule_from_file(self, update: Update, context: CallbackContext):
        """Получает файл Excel, считывает его, очищает базу и записывает данные."""
        file = update.message.document if update.message else None
        if file is None:
            await update.message.reply_text("Пожалуйста, прикрепите файл с расписанием.")
            return self.WAITING_FOR_FILE

        try:
            # Получаем объект File с использованием асинхронного вызова
            telegram_file = await file.get_file()
            file_path = await telegram_file.download_to_drive()  # Скачиваем файл

            # Открываем файл Excel и считываем данные
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            schedule_data = []
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
                if all(cell is None or cell == "" for cell in row):
                    continue  # Пропускаем пустые строки

                date, day_of_week, status, start_shift, start_break, end_break, end_shift = row

                # Пропускаем строки, которые явно не являются данными
                if not date or not day_of_week or not status:
                    continue

                # Преобразуем дату в формат YYYY-MM-DD
                try:
                    date_obj = datetime.strptime(date, "%Y-%m-%d")  # Исправленный формат
                    formatted_date = date_obj.strftime("%Y-%m-%d")
                except ValueError:
                    logger.error(f"Ошибка преобразования даты: {date}")
                    await update.message.reply_text(f"Ошибка в формате даты: {date}")
                    return ConversationHandler.END

                # Проверяем статус
                if status not in ["Рабочий", "Выходной"]:
                    logger.error(f"Некорректный статус: {status}")
                    await update.message.reply_text(f"Ошибка в статусе дня: {status}")
                    return ConversationHandler.END

                # Добавляем строку в список для загрузки в БД
                schedule_data.append((formatted_date, day_of_week, status, start_shift, start_break, end_break, end_shift))

            # Проверяем, есть ли валидные данные
            if not schedule_data:
                await update.message.reply_text("Файл не содержит корректных данных для загрузки.")
                return ConversationHandler.END

            logger.info(f"Файл успешно считан. Готовимся к обновлению базы данных...")

            # Очищаем базу **только после успешного прочтения файла**
            logger.info(f"Очищаем базу перед вставкой новых данных...")
            self.db.clear_schedule()

            # Вставляем новые данные в базу
            logger.info(f"Вставляем данные в базу: {schedule_data}")
            self.db.insert_schedule(schedule_data)

            await update.message.reply_text("Расписание успешно обновлено в базе данных.")

            return ConversationHandler.END  # Завершаем разговор

        except Exception as e:
            logger.error(f"Ошибка при обработке файла: {e}", exc_info=True)
            await update.message.reply_text("Произошла ошибка при обработке файла. База не изменена.")
            return ConversationHandler.END  # Завершаем разговор в случае ошибки

    async def cancel(self, update: Update, context: CallbackContext):
        """Отмена процесса получения файла."""
        await update.message.reply_text("Процесс был отменен.")
        return ConversationHandler.END
